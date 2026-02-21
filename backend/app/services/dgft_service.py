"""
Feature 1: DGFT-Ready Excel Generator (Revenue Recovery)
Generates validated .xlsx file matching official DGFT bulk eBRC upload template.
"""

import io
from typing import Optional, List, Dict, Any
from datetime import datetime, timezone
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
import logging

from ..core.database import db
from ..common.utils import generate_id, now_iso
from .tenant_auth_service import TenantAuthService

logger = logging.getLogger(__name__)

# DGFT eBRC Template Column Mapping
DGFT_COLUMNS = [
    "SL_NO", "SHIPPING_BILL_NO", "SHIPPING_BILL_DATE", "PORT_CODE",
    "IEC_CODE", "EXPORTER_NAME", "INVOICE_NO", "INVOICE_DATE",
    "INVOICE_VALUE", "INVOICE_CURRENCY", "FOB_VALUE_INR",
    "IRM_REF_NO", "IRM_DATE", "BANK_AD_CODE", "BANK_NAME",
    "REALIZED_AMOUNT", "REALIZATION_DATE", "HS_CODE",
    "PRODUCT_DESCRIPTION", "BUYER_NAME", "BUYER_COUNTRY"
]

# Cell styling
RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


class DGFTExcelService:
    """Service for generating DGFT-compliant Excel files for eBRC bulk upload."""
    
    @staticmethod
    async def get_shipments_with_bank_data(company_id: str, filters: Optional[Dict] = None) -> List[Dict]:
        """
        Performs async join between shipments and connectors collections.
        Pulls IRM Reference Number and Bank AD Code from bank remittance data.
        """
        query = {"company_id": company_id}
        if filters:
            if filters.get("status"):
                query["status"] = filters["status"]
            if filters.get("start_date"):
                query["created_at"] = {"$gte": filters["start_date"]}
            if filters.get("end_date"):
                query.setdefault("created_at", {})["$lte"] = filters["end_date"]
        
        shipments = await db.shipments.find(query, {"_id": 0}).to_list(1000)
        if not shipments:
            return []
        
        company = await db.companies.find_one({"id": company_id}, {"_id": 0})
        
        # Get bank connector data
        connectors = await db.connectors.find({
            "company_id": company_id,
            "connector_type": "bank"
        }, {"_id": 0}).to_list(100)
        
        shipment_ids = [s["id"] for s in shipments]
        documents = await db.documents.find({"shipment_id": {"$in": shipment_ids}}, {"_id": 0}).to_list(2000)
        payments = await db.payments.find({"shipment_id": {"$in": shipment_ids}}, {"_id": 0}).to_list(2000)
        
        # Create lookups
        docs_by_shipment = {}
        for doc in documents:
            sid = doc.get("shipment_id")
            docs_by_shipment.setdefault(sid, []).append(doc)
        
        payments_by_shipment = {}
        for payment in payments:
            sid = payment.get("shipment_id")
            payments_by_shipment.setdefault(sid, []).append(payment)
        
        # Bank remittance lookup
        bank_data = {}
        for conn in connectors:
            remittances = conn.get("remittances", [])
            for rem in remittances:
                ref_shipment = rem.get("shipment_id")
                if ref_shipment:
                    bank_data[ref_shipment] = {
                        "irm_ref_no": rem.get("irm_reference_number"),
                        "irm_date": rem.get("irm_date"),
                        "bank_ad_code": conn.get("ad_code") or rem.get("ad_code"),
                        "bank_name": conn.get("bank_name") or rem.get("bank_name"),
                        "realized_amount": rem.get("realized_amount"),
                        "realization_date": rem.get("realization_date")
                    }
        
        result = []
        for shipment in shipments:
            sid = shipment["id"]
            ship_docs = docs_by_shipment.get(sid, [])
            ship_payments = payments_by_shipment.get(sid, [])
            ship_bank = bank_data.get(sid, {})
            
            invoice_doc = next((d for d in ship_docs if d.get("document_type") == "invoice"), None)
            sb_doc = next((d for d in ship_docs if d.get("document_type") == "shipping_bill"), None)
            
            total_realized = ship_bank.get("realized_amount")
            if total_realized is None:
                total_realized = sum(p.get("amount", 0) for p in ship_payments 
                                   if p.get("status") in ["completed", "received", "paid"])
            
            realization_date = ship_bank.get("realization_date")
            if not realization_date and ship_payments:
                paid = [p for p in ship_payments if p.get("status") in ["completed", "received", "paid"]]
                if paid:
                    realization_date = max(p.get("received_date") or p.get("created_at") for p in paid)
            
            result.append({
                "shipment": shipment,
                "company": company,
                "invoice": invoice_doc,
                "shipping_bill": sb_doc,
                "bank_data": ship_bank,
                "payments": ship_payments,
                "realized_amount": total_realized,
                "realization_date": realization_date
            })
        
        return result
    
    @staticmethod
    def format_date(date_str: Optional[str]) -> str:
        """Convert ISO date to DGFT format (DD/MM/YYYY)."""
        if not date_str:
            return ""
        try:
            dt = datetime.fromisoformat(str(date_str).replace("Z", "+00:00"))
            return dt.strftime("%d/%m/%Y")
        except:
            return str(date_str) if date_str else ""
    
    @staticmethod
    async def generate_dgft_excel(company_id: str, user_id: str, filters: Optional[Dict] = None) -> StreamingResponse:
        """
        Generate DGFT-compliant Excel file for eBRC bulk upload.
        Missing data cells are highlighted in red.
        """
        data = await DGFTExcelService.get_shipments_with_bank_data(company_id, filters)
        
        if not data:
            raise HTTPException(status_code=404, detail="No shipments found for export")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "DGFT_eBRC_Upload"
        
        # Add header row
        for col_idx, col_name in enumerate(DGFT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
        
        missing_summary = {col: 0 for col in DGFT_COLUMNS}
        
        for row_idx, record in enumerate(data, 2):
            shipment = record["shipment"]
            company = record["company"] or {}
            invoice = record["invoice"]
            sb = record["shipping_bill"]
            bank = record["bank_data"]
            
            inv_data = (invoice.get("data", {}) or invoice.get("extracted_data", {})) if invoice else {}
            sb_data = (sb.get("data", {}) or sb.get("extracted_data", {})) if sb else {}
            
            row_data = {
                "SL_NO": row_idx - 1,
                "SHIPPING_BILL_NO": sb_data.get("sb_number") or (sb.get("document_number") if sb else None),
                "SHIPPING_BILL_DATE": DGFTExcelService.format_date(sb_data.get("sb_date") or shipment.get("actual_ship_date")),
                "PORT_CODE": sb_data.get("port_code") or shipment.get("origin_port"),
                "IEC_CODE": company.get("iec_code") or company.get("iec"),
                "EXPORTER_NAME": company.get("name") or company.get("company_name"),
                "INVOICE_NO": inv_data.get("invoice_number") or (invoice.get("document_number") if invoice else None),
                "INVOICE_DATE": DGFTExcelService.format_date(inv_data.get("invoice_date")),
                "INVOICE_VALUE": inv_data.get("total_amount") or shipment.get("total_value"),
                "INVOICE_CURRENCY": shipment.get("currency", "USD"),
                "FOB_VALUE_INR": shipment.get("fob_value_inr") or (shipment.get("total_value", 0) * 83.5),
                "IRM_REF_NO": bank.get("irm_ref_no"),
                "IRM_DATE": DGFTExcelService.format_date(bank.get("irm_date")),
                "BANK_AD_CODE": bank.get("bank_ad_code"),
                "BANK_NAME": bank.get("bank_name"),
                "REALIZED_AMOUNT": record["realized_amount"],
                "REALIZATION_DATE": DGFTExcelService.format_date(record["realization_date"]),
                "HS_CODE": shipment.get("hs_codes", [None])[0] if shipment.get("hs_codes") else None,
                "PRODUCT_DESCRIPTION": shipment.get("product_description"),
                "BUYER_NAME": shipment.get("buyer_name"),
                "BUYER_COUNTRY": shipment.get("buyer_country")
            }
            
            required_fields = ["SHIPPING_BILL_NO", "SHIPPING_BILL_DATE", "IEC_CODE", "INVOICE_NO", "IRM_REF_NO", "BANK_AD_CODE"]
            warning_fields = ["INVOICE_DATE", "REALIZED_AMOUNT", "REALIZATION_DATE"]
            
            for col_idx, col_name in enumerate(DGFT_COLUMNS, 1):
                value = row_data.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                
                if col_name in required_fields and (value is None or value == ""):
                    cell.fill = RED_FILL
                    missing_summary[col_name] += 1
                elif col_name in warning_fields and (value is None or value == ""):
                    cell.fill = YELLOW_FILL
                    missing_summary[col_name] += 1
                
                if col_name in ["INVOICE_VALUE", "FOB_VALUE_INR", "REALIZED_AMOUNT"] and value:
                    cell.number_format = '#,##0.00'
        
        # Auto-adjust column widths
        for col_idx, col_name in enumerate(DGFT_COLUMNS, 1):
            max_len = len(col_name)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 30)
        
        # Add summary sheet
        summary_ws = wb.create_sheet("Data_Quality_Summary")
        summary_ws.cell(row=1, column=1, value="Field").font = HEADER_FONT
        summary_ws.cell(row=1, column=1).fill = HEADER_FILL
        summary_ws.cell(row=1, column=2, value="Missing Count").font = HEADER_FONT
        summary_ws.cell(row=1, column=2).fill = HEADER_FILL
        summary_ws.cell(row=1, column=3, value="Status").font = HEADER_FONT
        summary_ws.cell(row=1, column=3).fill = HEADER_FILL
        
        for row_idx, (field, count) in enumerate(missing_summary.items(), 2):
            summary_ws.cell(row=row_idx, column=1, value=field)
            summary_ws.cell(row=row_idx, column=2, value=count)
            status = "OK" if count == 0 else "ACTION REQUIRED" if field in required_fields else "WARNING"
            summary_ws.cell(row=row_idx, column=3, value=status)
            if count > 0:
                summary_ws.cell(row=row_idx, column=3).fill = RED_FILL if status == "ACTION REQUIRED" else YELLOW_FILL
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        filename = f"DGFT_eBRC_Export_{timestamp}.xlsx"
        
        # Audit log
        await db.audit_logs.insert_one({
            "id": generate_id(),
            "action_type": "dgft_excel_export",
            "resource_type": "dgft_export",
            "company_id": company_id,
            "user_id": user_id,
            "details": {
                "record_count": len(data),
                "missing_data_summary": {k: v for k, v in missing_summary.items() if v > 0}
            },
            "timestamp": now_iso()
        })
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "X-Record-Count": str(len(data)),
                "X-Data-Quality-Issues": str(sum(1 for v in missing_summary.values() if v > 0))
            }
        )
    
    @staticmethod
    async def validate_dgft_data(company_id: str) -> Dict[str, Any]:
        """Validate shipment data against DGFT schema requirements."""
        data = await DGFTExcelService.get_shipments_with_bank_data(company_id)
        
        errors = []
        warnings = []
        
        for idx, record in enumerate(data):
            shipment = record["shipment"]
            bank = record["bank_data"]
            row_num = idx + 1
            
            if not bank.get("irm_ref_no"):
                errors.append({
                    "row": row_num,
                    "shipment_id": shipment["id"],
                    "shipment_number": shipment.get("shipment_number"),
                    "field": "IRM_REF_NO",
                    "message": "Missing IRM Reference Number - connect bank via Account Aggregator"
                })
            
            if not bank.get("bank_ad_code"):
                errors.append({
                    "row": row_num,
                    "shipment_id": shipment["id"],
                    "field": "BANK_AD_CODE",
                    "message": "Missing Bank AD Code - required for DGFT submission"
                })
            
            if not record["realized_amount"]:
                warnings.append({
                    "row": row_num,
                    "shipment_id": shipment["id"],
                    "field": "REALIZED_AMOUNT",
                    "message": "No realization recorded - may be pending payment"
                })
        
        return {
            "total_records": len(data),
            "valid_records": len(data) - len(errors),
            "error_count": len(errors),
            "warning_count": len(warnings),
            "errors": errors,
            "warnings": warnings,
            "is_valid": len(errors) == 0
        }
